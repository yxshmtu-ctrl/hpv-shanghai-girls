#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""gen_class_list.py — 花名册 → 免费适龄名单(Excel)

输入：学生花名册 .xlsx/.csv，自动识别列名：
  姓名/名字/学生姓名、性别/性别/男/女、出生日期/出生年月/生日/出生日、班级/班别/班
规则：2011-11-10 之后出生 + 满 13 周岁 + 女 → 免费适龄（复用 check_eligibility）
输出：.xlsx 名单，免费适龄者高亮；含状态/年龄/预计符合日期

用法：
  python scripts/gen_class_list.py 花名册.xlsx --out 适龄名单.xlsx
  python scripts/gen_class_list.py 花名册.xlsx --class 六年级(1)班
  python scripts/gen_class_list.py 花名册.csv --today 2026-09-01   # 测试指定日期

⚠️ 个人信息合规：花名册含未成年人个人信息，仅限授权/加密环境使用，禁止上传公开仓库。
"""

import sys, os, argparse, re

sys.stdout.reconfigure(encoding="utf-8")

try:
    import pandas as pd
except ImportError:
    sys.exit("需要 pandas：pip install pandas openpyxl")

# 复用同目录判定逻辑
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from check_eligibility import check_eligibility, parse_date  # noqa: E402

SEX_FEMALE = {"女", "f", "female", "2"}
SEX_MALE = {"男", "m", "male", "1"}

COL_ALIASES = {
    "name": ["姓名", "名字", "学生姓名", "name"],
    "sex": ["性别", "sex", "gender"],
    "birth": ["出生日期", "出生年月", "生日", "出生日", "birth", "birthday", "出生"],
    "class": ["班级", "班别", "班", "class", "classname", "年级班级"],
}


def resolve_col(df, field):
    cols = {str(c).strip() for c in df.columns}
    low = {str(c).strip().lower() for c in df.columns}
    for alias in COL_ALIASES[field]:
        if alias in cols or alias.lower() in low:
            for c in df.columns:
                if str(c).strip().lower() == alias.lower() or str(c).strip() == alias:
                    return c
    return None


def parse_birth(v):
    v = str(v).strip()
    if not v or v.lower() in {"nan", "none", "unknown", "-"}:
        return None
    # 容忍 2013/5/1、2013-5-1、20130501、2013.5
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
        try:
            return parse_date(pd.to_datetime(v, format=fmt).strftime("%Y-%m-%d"))
        except Exception:
            pass
    try:
        s = pd.to_datetime(v)
        return parse_date(s.strftime("%Y-%m-%d"))
    except Exception:
        return None


def main():
    ap = argparse.ArgumentParser(description="花名册 → 免费适龄名单")
    ap.add_argument("roster", help="花名册 .xlsx/.csv")
    ap.add_argument("--out", default=None, help="输出 xlsx 路径，默认 适龄名单.xlsx")
    ap.add_argument("--class", dest="cls", default=None, help="只处理某班级")
    ap.add_argument("--today", default=None, help="判定日期 YYYY-MM-DD（测试用）")
    a = ap.parse_args()

    if a.roster.lower().endswith(".csv"):
        df = pd.read_csv(a.roster, dtype=str)
    else:
        df = pd.read_excel(a.roster, dtype=str)

    c_name = resolve_col(df, "name")
    c_sex = resolve_col(df, "sex")
    c_birth = resolve_col(df, "birth")
    c_class = resolve_col(df, "class")

    if not c_name or not c_birth:
        sys.exit(
            f"未识别到必要列。现有列：{list(df.columns)}\n需含 姓名 与 出生日期 列。"
        )

    if a.cls:
        if not c_class:
            sys.exit("指定了 --class，但花名册没有班级列。")
        df = df[df[c_class].astype(str).str.contains(a.cls, na=False, regex=False)]

    rows = []
    for _, r in df.iterrows():
        name = str(r[c_name]).strip() if c_name and pd.notna(r[c_name]) else "?"
        bd = parse_birth(r[c_birth]) if c_birth and pd.notna(r[c_birth]) else None
        sex = str(r[c_sex]).strip().lower() if c_sex and pd.notna(r[c_sex]) else ""
        cls = str(r[c_class]).strip() if c_class and pd.notna(r[c_class]) else ""

        if sex in SEX_MALE:
            status = "非女生（可自费）"
            eligible, age, est = False, None, None
        elif bd is None:
            status = "出生日期无法解析"
            eligible, age, est = False, None, None
        else:
            res = check_eligibility(
                bd.isoformat(), parse_date(a.today) if a.today else None
            )
            eligible = res["eligible"]
            age = res["age"]
            est = res["estimated_eligible_date"]
            if eligible:
                status = "免费适龄"
            elif est:
                status = f"未满13周岁（预计{est}符合）"
            elif not res["birth_after_ok"]:
                status = "超出生日期范围（可自费）"
            else:
                status = "不符合（可自费）"

        rows.append(
            {
                "姓名": name,
                "班级": cls,
                "出生日期": bd.isoformat() if bd else "",
                "现年龄": age if age is not None else "",
                "是否符合免费": "是" if eligible else "否",
                "状态": status,
            }
        )

    out_df = pd.DataFrame(rows)
    out = a.out or "适龄名单.xlsx"
    os.makedirs(os.path.dirname(os.path.abspath(out)) or ".", exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        out_df.to_excel(w, index=False, sheet_name="适龄判定")
        # 免费适龄者高亮
        wb = w.book
        ws = wb["适龄判定"]
        from openpyxl.styles import PatternFill

        green = PatternFill("solid", fgColor="C6EFCE")
        col_elig = list(out_df.columns).index("是否符合免费") + 1
        for i, v in enumerate(out_df["是否符合免费"], start=2):
            if v == "是":
                for j in range(1, len(out_df.columns) + 1):
                    ws.cell(row=i, column=j).fill = green
    print(
        f"已生成：{out}  （共{len(out_df)}人，免费适龄{int((out_df['是否符合免费'] == '是').sum())}人）"
    )


if __name__ == "__main__":
    main()
